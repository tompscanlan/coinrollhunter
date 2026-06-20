#!/usr/bin/env python3
"""Coins & bullion portfolio engine.

Combines pm_holdings.json (bullion + silver finds) with crh_ledger.json
(coin-roll-hunting operating costs) to answer:

    Is the bullion investment profitable, and is coin roll hunting
    paying for itself - and have we cashed in everything not kept?

Two activities are reported separately:
  * BULLION  - gold + purchased junk silver. A long-term investment.
  * CRH      - silver/keeper finds minus the cost of the hunt.

Core math is stdlib-only. --xlsx needs openpyxl; --html is stdlib.

Usage:
    python portfolio.py                          # console report, JSON spot
    python portfolio.py --gold 4167 --silver 65  # override spot
    python portfolio.py --xlsx Coins_Portfolio.xlsx --html dashboard.html
"""
from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path

HERE = Path(__file__).resolve().parent


# ----------------------------- load + compute -----------------------------
def load(path: Path) -> dict:
    with path.open(encoding="utf-8") as fh:
        return json.load(fh)


def is_find(lot: dict) -> bool:
    return str(lot.get("id", "")).upper().startswith("FIND")


def enrich(lot: dict, spot: dict) -> dict:
    fine_oz = lot["qty"] * lot["fine_oz_each"]
    market = fine_oz * spot[lot["metal"]]
    basis = lot["basis_usd"]
    unreal = market - basis
    pct = (unreal / basis * 100) if basis else float("inf")
    return {**lot, "fine_oz": fine_oz, "market_usd": market,
            "unreal_usd": unreal, "unreal_pct": pct}


def buyback_factor(lot: dict, s: dict) -> float:
    """Estimated realizable dealer payout vs melt for junk silver."""
    if lot["metal"] != "silver":
        return 1.0
    fin = str(lot.get("fineness", ""))
    if fin.startswith("40"):
        return s.get("silver_buyback_factor_40pct", 0.80)
    if fin.startswith("90"):
        return s.get("silver_buyback_factor_90pct", 0.90)
    return 1.0


def compute(holdings: dict, crh: dict, spot: dict) -> dict:
    lots = [enrich(l, spot) for l in holdings["lots"]]
    s = crh.get("settings", {})

    bullion = [l for l in lots if not is_find(l)]
    finds = [l for l in lots if is_find(l)]

    def total(rows, key):
        return sum(r[key] for r in rows)

    # --- Bullion investment ---
    b_basis = total(bullion, "basis_usd")
    b_market = total(bullion, "market_usd")
    b_unreal = b_market - b_basis

    gold = [l for l in bullion if l["metal"] == "gold"]
    g_oz, g_basis, g_market = total(gold, "fine_oz"), total(gold, "basis_usd"), total(gold, "market_usd")

    # --- CRH finds ---
    f_cost = total(finds, "basis_usd")           # face paid
    f_melt = total(finds, "market_usd")          # full melt at spot
    f_realizable = sum(l["market_usd"] * buyback_factor(l, s) for l in finds)
    f_oz = total(finds, "fine_oz")

    # --- CRH operating costs ---
    mileage_rate = s.get("irs_mileage_rate_usd_per_mile", 0.70)
    trips = [t for t in crh.get("trips", []) if "miles" in t]
    gas = sum(t.get("miles", 0) * mileage_rate for t in trips)
    hours = sum(t.get("hours", 0) for t in trips)
    supplies = sum(x.get("cost_usd", 0) for x in crh.get("supplies", []) if "cost_usd" in x)
    op_cost = gas + supplies

    # --- float, kept, cash-in reconciliation ---
    rolls = crh.get("roll_transactions", [])
    buys = sum(t["cash_usd"] for t in rolls if t["action"] == "buy")
    returns = sum(t["cash_usd"] for t in rolls if t["action"] == "return")
    kc = crh.get("keepers_clad", {})
    clad_face = kc.get("halves_face_usd", 0) + kc.get("quarters_face_usd", 0)
    kept_face = clad_face + f_cost
    to_redeposit = buys - returns - kept_face   # face not kept and not yet cashed in
    float_out = to_redeposit
    reconciled = abs(to_redeposit) < 0.01

    # --- box throughput ---
    boxes_by_denom = {}
    for t in rolls:
        if t["action"] == "buy" and t.get("boxes"):
            boxes_by_denom[t["denom"]] = boxes_by_denom.get(t["denom"], 0) + t["boxes"]
    total_boxes = sum(boxes_by_denom.values())
    face_searched = buys

    # --- CRH net ---
    crh_net_melt = f_melt - f_cost - op_cost
    crh_net_real = f_realizable - f_cost - op_cost
    rate = s.get("hourly_rate_usd", 0)
    crh_net_time = crh_net_real - hours * rate

    # --- totals ---
    t_basis = b_basis + f_cost
    t_market = b_market + f_realizable
    t_unreal = t_market - t_basis

    return {
        "as_of": holdings.get("spot_reference", {}).get("as_of", str(date.today())),
        "spot": spot,
        "lots": lots, "bullion": bullion, "finds": finds,
        "bullion_basis": b_basis, "bullion_market": b_market, "bullion_unreal": b_unreal,
        "bullion_pct": (b_unreal / b_basis * 100) if b_basis else 0,
        "gold_oz": g_oz, "gold_basis": g_basis, "gold_market": g_market,
        "find_cost": f_cost, "find_melt": f_melt, "find_realizable": f_realizable, "find_oz": f_oz,
        "gas": gas, "hours": hours, "supplies": supplies, "op_cost": op_cost,
        "buys": buys, "returns": returns, "clad_face": clad_face, "float_out": float_out,
        "kept_face": kept_face, "to_redeposit": to_redeposit, "reconciled": reconciled,
        "boxes_by_denom": boxes_by_denom, "total_boxes": total_boxes, "face_searched": face_searched,
        "crh_net_melt": crh_net_melt, "crh_net_real": crh_net_real, "crh_net_time": crh_net_time,
        "hourly_rate": rate, "value_time": s.get("value_time", False),
        "total_basis": t_basis, "total_market": t_market, "total_unreal": t_unreal,
        "crh": crh,
    }


# ----------------------------- formatting -----------------------------
def m(x):
    return f"${x:,.2f}"


def p(x):
    return "n/a" if x == float("inf") else f"{x:+.1f}%"


def verdict(r: dict) -> str:
    if r["crh_net_real"] > 0:
        return "PROFITABLE (cash basis)"
    if r["crh_net_real"] == 0:
        return "BREAK-EVEN"
    return "COSTING MONEY"


def console_report(r: dict) -> str:
    L = []
    L.append("=" * 64)
    L.append(f"  COINS & BULLION PORTFOLIO  -  as of {r['as_of']}")
    L.append(f"  Spot: Au {m(r['spot']['gold'])} / Ag {m(r['spot']['silver'])} per ozt")
    L.append("=" * 64)
    L.append("")
    L.append("BULLION INVESTMENT (gold + purchased junk silver)")
    L.append(f"  Gold:    {r['gold_oz']:.4f} oz   basis {m(r['gold_basis'])}   market {m(r['gold_market'])}")
    L.append(f"  Basis {m(r['bullion_basis'])}  Market {m(r['bullion_market'])}  "
             f"Unrealized {m(r['bullion_unreal'])} ({p(r['bullion_pct'])})")
    L.append("")
    L.append("COIN ROLL HUNTING (finds minus cost of the hunt)")
    L.append(f"  Silver found:    {r['find_oz']:.4f} oz fine")
    L.append(f"  Face cost paid:  {m(r['find_cost'])}")
    L.append(f"  Melt value:      {m(r['find_melt'])}   (realizable ~{m(r['find_realizable'])} after dealer haircut)")
    L.append(f"  Gas (logged):    {m(r['gas'])}    Supplies: {m(r['supplies'])}")
    L.append(f"  Clad keepers parked at face: {m(r['clad_face'])} (recoverable, not a loss)")
    L.append("")
    boxes = ", ".join(f"{v:g} {k}" for k, v in r['boxes_by_denom'].items()) or "none logged"
    L.append(f"  Boxes searched: {r['total_boxes']:g} total ({boxes}) = {m(r['face_searched'])} face")
    L.append("  Cash-in check  (have we redeposited everything not kept?)")
    L.append(f"    Bought {m(r['buys'])}  -  Returned {m(r['returns'])}  -  Kept {m(r['kept_face'])}")
    if r['reconciled']:
        L.append("    => $0.00 outstanding. ALL CASHED IN.")
    else:
        L.append(f"    => {m(r['to_redeposit'])} STILL TO REDEPOSIT (searched culls / coin left to search)")
    L.append("  " + "-" * 50)
    L.append(f"  CRH NET (realizable, cash only): {m(r['crh_net_real'])}   -> {verdict(r)}")
    L.append(f"  CRH NET (full melt):            {m(r['crh_net_melt'])}")
    if r["hourly_rate"]:
        L.append(f"  CRH NET if time valued @ {m(r['hourly_rate'])}/hr ({r['hours']:.1f} h logged): {m(r['crh_net_time'])}")
    L.append("")
    L.append("WHOLE PORTFOLIO")
    L.append(f"  Cash invested: {m(r['total_basis'])}   Liquidation value (est): {m(r['total_market'])}")
    L.append(f"  Net position:  {m(r['total_unreal'])} ({p(r['total_unreal']/r['total_basis']*100)})")
    L.append("=" * 64)
    return "\n".join(L)


# ----------------------------- xlsx -----------------------------
def write_xlsx(r: dict, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY = "1F3864"; GREEN = "C6EFCE"; RED = "FFC7CE"
    GREENF = "006100"; REDF = "9C0006"
    hdr = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor=NAVY)
    title = Font(bold=True, size=15, color=NAVY)
    sub = Font(italic=True, size=9, color="666666")
    bold = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money = '$#,##0.00'; pct = '+0.0%;-0.0%'; oz = '0.0000'

    wb = Workbook()

    def style_header(ws, row, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr; cell.fill = hfill; cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---------- Summary ----------
    ws = wb.active; ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Coins & Bullion - Profitability"; ws["A1"].font = title
    ws["A2"] = f"As of {r['as_of']}  |  Spot Au {m(r['spot']['gold'])} / Ag {m(r['spot']['silver'])} per ozt"
    ws["A2"].font = sub

    def block(ws, top, head, rows):
        ws.cell(row=top, column=1, value=head).font = Font(bold=True, size=12, color=NAVY)
        rr = top + 1
        for label, val, fmt, *flag in rows:
            ws.cell(row=rr, column=1, value=label)
            c = ws.cell(row=rr, column=2, value=val)
            if fmt:
                c.number_format = fmt
            if flag and flag[0] == "v":
                c.font = bold
                if isinstance(val, (int, float)):
                    c.fill = PatternFill("solid", fgColor=GREEN if val >= 0 else RED)
                    c.font = Font(bold=True, color=GREENF if val >= 0 else REDF)
            rr += 1
        return rr + 1

    nxt = block(ws, 4, "BULLION INVESTMENT", [
        ("Gold held (oz)", r["gold_oz"], oz),
        ("Cash basis", r["bullion_basis"], money),
        ("Market value", r["bullion_market"], money),
        ("Unrealized P/L", r["bullion_unreal"], money, "v"),
        ("Return %", r["bullion_pct"] / 100, pct, "v"),
    ])
    nxt = block(ws, nxt, "COIN ROLL HUNTING", [
        ("Silver found (oz fine)", r["find_oz"], oz),
        ("Face cost paid", r["find_cost"], money),
        ("Melt value", r["find_melt"], money),
        ("Realizable (after haircut)", r["find_realizable"], money),
        ("Gas + supplies (logged)", r["op_cost"], money),
        ("CRH NET (cash, realizable)", r["crh_net_real"], money, "v"),
        ("CRH NET if time valued", r["crh_net_time"], money, "v"),
        ("Clad keepers parked @ face", r["clad_face"], money),
        ("Boxes searched (total)", r["total_boxes"], '0.0#'),
        ("Still to redeposit", r["to_redeposit"], money),
    ])
    nxt = block(ws, nxt, "WHOLE PORTFOLIO", [
        ("Total cash invested", r["total_basis"], money),
        ("Liquidation value (est)", r["total_market"], money),
        ("Net position", r["total_unreal"], money, "v"),
    ])
    ws.cell(row=nxt, column=1, value=f"CRH verdict: {verdict(r)}").font = Font(bold=True, size=12,
        color=GREENF if r["crh_net_real"] >= 0 else REDF)
    msg = "All searched coin not kept has been cashed in." if r["reconciled"] else \
        f"{m(r['to_redeposit'])} of searched/unsearched coin still to redeposit."
    ws.cell(row=nxt + 1, column=1, value="Cash-in: " + msg).font = sub
    ws.cell(row=nxt + 2, column=1,
            value="Bullion = long-term investment (mark-to-market). CRH should at least pay for itself.").font = sub
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 18

    # ---------- Bullion lots ----------
    ws2 = wb.create_sheet("Bullion")
    ws2.sheet_view.showGridLines = False
    cols = ["Lot", "Product", "Acquired", "Metal", "Fine oz", "Basis $", "Market $", "Unreal $", "%"]
    for i, c in enumerate(cols, 1):
        ws2.cell(row=1, column=i, value=c)
    style_header(ws2, 1, len(cols))
    rr = 2
    for l in r["bullion"]:
        vals = [l["id"], l["product"], l["acquired"], l["metal"], l["fine_oz"],
                l["basis_usd"], l["market_usd"], l["unreal_usd"], l["unreal_pct"] / 100]
        for i, v in enumerate(vals, 1):
            cell = ws2.cell(row=rr, column=i, value=v); cell.border = border
            if i == 5: cell.number_format = oz
            if i in (6, 7, 8): cell.number_format = money
            if i == 9: cell.number_format = pct
            if i == 8: cell.font = Font(color=GREENF if l["unreal_usd"] >= 0 else REDF)
        rr += 1
    ws2.cell(row=rr, column=1, value="TOTAL").font = bold
    ws2.cell(row=rr, column=5, value=sum(l["fine_oz"] for l in r["bullion"])).number_format = oz
    ws2.cell(row=rr, column=6, value=r["bullion_basis"]).number_format = money
    ws2.cell(row=rr, column=7, value=r["bullion_market"]).number_format = money
    ws2.cell(row=rr, column=8, value=r["bullion_unreal"]).number_format = money
    for c in range(1, 9): ws2.cell(row=rr, column=c).font = bold
    widths = [16, 40, 12, 8, 10, 12, 12, 12, 9]
    for i, w in enumerate(widths, 1): ws2.column_dimensions[get_column_letter(i)].width = w

    # ---------- CRH detail ----------
    ws3 = wb.create_sheet("CRH")
    ws3.sheet_view.showGridLines = False
    ws3["A1"] = "Coin Roll Hunting - finds, costs, boxes, float"; ws3["A1"].font = title
    # Finds table
    ws3["A3"] = "FINDS (silver pulled from rolls)"; ws3["A3"].font = Font(bold=True, color=NAVY)
    fcols = ["Lot", "Product", "Found", "Fine oz", "Face cost $", "Melt $", "Realizable $"]
    for i, c in enumerate(fcols, 1): ws3.cell(row=4, column=i, value=c)
    style_header(ws3, 4, len(fcols))
    s = r["crh"].get("settings", {})
    rr = 5
    for l in r["finds"]:
        bf = buyback_factor(l, s)
        vals = [l["id"], l["product"], l["acquired"], l["fine_oz"], l["basis_usd"],
                l["market_usd"], l["market_usd"] * bf]
        for i, v in enumerate(vals, 1):
            cell = ws3.cell(row=rr, column=i, value=v); cell.border = border
            if i == 4: cell.number_format = oz
            if i in (5, 6, 7): cell.number_format = money
        rr += 1
    ws3.cell(row=rr, column=1, value="TOTAL").font = bold
    ws3.cell(row=rr, column=4, value=r["find_oz"]).number_format = oz
    ws3.cell(row=rr, column=5, value=r["find_cost"]).number_format = money
    ws3.cell(row=rr, column=6, value=r["find_melt"]).number_format = money
    ws3.cell(row=rr, column=7, value=r["find_realizable"]).number_format = money
    for c in range(1, 8): ws3.cell(row=rr, column=c).font = bold

    # Cash-in reconciliation
    base = rr + 3
    ws3.cell(row=base - 1, column=1, value="CASH-IN RECONCILIATION").font = Font(bold=True, color=NAVY)
    recon = [("Bought (face)", r["buys"]), ("Returned (face)", r["returns"]),
             ("Kept: finds + clad", r["kept_face"]), ("Still to redeposit", r["to_redeposit"])]
    rr = base
    for lab, val in recon:
        ws3.cell(row=rr, column=1, value=lab)
        ws3.cell(row=rr, column=2, value=val).number_format = money
        rr += 1
    ws3.cell(row=rr, column=1, value=("ALL CASHED IN" if r["reconciled"] else "OUTSTANDING - redeposit due")).font = \
        Font(bold=True, color=GREENF if r["reconciled"] else REDF)

    # Boxes searched
    bb = rr + 2
    ws3.cell(row=bb - 1, column=1, value="BOXES SEARCHED (throughput)").font = Font(bold=True, color=NAVY)
    for i, c in enumerate(["Denomination", "Boxes"], 1): ws3.cell(row=bb, column=i, value=c)
    style_header(ws3, bb, 2)
    rr = bb + 1
    for den, n in r["boxes_by_denom"].items():
        ws3.cell(row=rr, column=1, value=den); ws3.cell(row=rr, column=2, value=n); rr += 1
    ws3.cell(row=rr, column=1, value="TOTAL").font = bold
    ws3.cell(row=rr, column=2, value=r["total_boxes"]).font = bold

    # Roll transactions
    base = rr + 3
    ws3.cell(row=base - 1, column=1, value="ROLL TRANSACTIONS (cash float)").font = Font(bold=True, color=NAVY)
    tcols = ["Date", "Bank", "Action", "Cash $", "Denom", "Boxes", "Notes"]
    for i, c in enumerate(tcols, 1): ws3.cell(row=base, column=i, value=c)
    style_header(ws3, base, len(tcols))
    rr = base + 1
    for t in r["crh"].get("roll_transactions", []):
        vals = [t["date"], t["bank"], t["action"], t["cash_usd"], t.get("denom", ""), t.get("boxes", ""), t.get("notes", "")]
        for i, v in enumerate(vals, 1):
            cell = ws3.cell(row=rr, column=i, value=v); cell.border = border
            if i == 4: cell.number_format = money
        rr += 1

    # Trip / gas log
    tb = rr + 3
    ws3.cell(row=tb - 1, column=1, value="TRIP / GAS LOG  (add rows: date, bank, round-trip miles, hours)").font = Font(bold=True, color=NAVY)
    for i, c in enumerate(["Date", "Bank", "Miles", "Hours", "Gas $ (auto)"], 1):
        ws3.cell(row=tb, column=i, value=c)
    style_header(ws3, tb, 5)
    rate = s.get("irs_mileage_rate_usd_per_mile", 0.70)
    rr = tb + 1
    for t in r["crh"].get("trips", []):
        if "miles" not in t: continue
        ws3.cell(row=rr, column=1, value=t.get("date"))
        ws3.cell(row=rr, column=2, value=t.get("bank"))
        ws3.cell(row=rr, column=3, value=t.get("miles"))
        ws3.cell(row=rr, column=4, value=t.get("hours", 0))
        ws3.cell(row=rr, column=5, value=f"=C{rr}*{rate}").number_format = money
        rr += 1
    widths = [16, 22, 12, 12, 14, 10, 30]
    for i, w in enumerate(widths, 1): ws3.column_dimensions[get_column_letter(i)].width = w

    # ---------- Tax lots ----------
    ws4 = wb.create_sheet("Tax Lots")
    ws4.sheet_view.showGridLines = False
    ws4["A1"] = "Cost-basis ledger (collectibles - max 28% LT rate)"; ws4["A1"].font = title
    ws4["A2"] = "Physical gold/silver = IRS 'collectibles'. >1yr held taxed up to 28%; <1yr at ordinary rates."
    ws4["A2"].font = sub
    tcols = ["Lot", "Product", "Acquired", "Basis $", "Holding", "LT? (>1yr)", "Status"]
    for i, c in enumerate(tcols, 1): ws4.cell(row=4, column=i, value=c)
    style_header(ws4, 4, len(tcols))
    today = date.today()
    rr = 5
    for l in r["lots"]:
        acq = datetime.strptime(l["acquired"], "%Y-%m-%d").date()
        held_days = (today - acq).days
        lt = held_days >= 365
        vals = [l["id"], l["product"], l["acquired"], l["basis_usd"],
                f"{held_days} days", "Yes" if lt else "No", "Held (unrealized)"]
        for i, v in enumerate(vals, 1):
            cell = ws4.cell(row=rr, column=i, value=v); cell.border = border
            if i == 4: cell.number_format = money
        rr += 1
    ws4.cell(row=rr + 1, column=1, value="No sales yet - all positions open. Realized gains logged here on disposal.").font = sub
    widths = [16, 40, 12, 12, 12, 12, 18]
    for i, w in enumerate(widths, 1): ws4.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)


# ----------------------------- html (static fallback) -----------------------------
def write_html(r: dict, path: Path) -> None:
    pos = r["crh_net_real"] >= 0
    vcolor = "#1a7d3c" if pos else "#b3261e"
    recon = "All searched coin not kept is cashed in." if r["reconciled"] else \
        f"{m(r['to_redeposit'])} still to redeposit."
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>Coins & Bullion (static)</title></head><body style="font-family:sans-serif;max-width:700px;margin:auto;padding:20px">
<h1>Coins &amp; Bullion - {r['as_of']}</h1>
<p>Spot Au {m(r['spot']['gold'])} / Ag {m(r['spot']['silver'])}</p>
<h2 style="color:{vcolor}">CRH: {verdict(r)} {m(r['crh_net_real'])}</h2>
<p>Bullion unrealized: {m(r['bullion_unreal'])} ({p(r['bullion_pct'])})</p>
<p>Boxes searched: {r['total_boxes']:g} | Cash-in: {recon}</p>
<p style="color:#888">This is a static snapshot. For live entry use the interactive dashboard.html.</p>
</body></html>"""
    path.write_text(html, encoding="utf-8")


# ----------------------------- main -----------------------------
def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--holdings", default=str(HERE / "pm_holdings.json"))
    ap.add_argument("--crh", default=str(HERE / "crh_ledger.json"))
    ap.add_argument("--gold", type=float, default=None)
    ap.add_argument("--silver", type=float, default=None)
    ap.add_argument("--xlsx", default=None, help="write Excel workbook to this path")
    ap.add_argument("--html", default=None, help="write static HTML snapshot to this path")
    args = ap.parse_args(argv)

    holdings = load(Path(args.holdings))
    crh = load(Path(args.crh))
    ref = holdings.get("spot_reference", {})
    spot = {
        "gold": args.gold if args.gold is not None else ref.get("gold_usd_per_ozt"),
        "silver": args.silver if args.silver is not None else ref.get("silver_usd_per_ozt"),
    }
    spot.update(crh.get("settings", {}))

    r = compute(holdings, crh, spot)
    print(console_report(r))
    if args.xlsx:
        write_xlsx(r, Path(args.xlsx)); print(f"\nWrote {args.xlsx}")
    if args.html:
        write_html(r, Path(args.html)); print(f"Wrote {args.html}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
